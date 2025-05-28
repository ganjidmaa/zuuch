import io
import json
import xlsxwriter
from openpyxl import load_workbook
from odoo import models
from odoo.tools import date_utils
from odoo import models, fields, api, _
from dateutil.relativedelta import relativedelta
import logging
from datetime import datetime, date



field_names = ['customer_type', 'customer_surname', 'customer_name', 'customer_registerno', 'customer_phone', 
               'state_number', 'insurance_name', 'contract_number', 'user_name', 'product_name', 'valuation', 
               'paid', 'payment_fee_percent', 'start_date', 'end_date', 'create_date', 'state', 'insurance_type_name']

field_names_1 = ['create_date', 'user_name', 'contract_number', 'paid', 'is_limit']

state_names = [
    {'state': 'draft', 'name': 'Ноорог'},
    {'state': 'paid', 'name': 'Төлөгдсөн'},
    {'state': 'sent', 'name': 'ERP Гэрээ бичсэн'}
]
customer_types = [
    {'state': 'person', 'name': 'Хувь хүн'},
    {'state': 'company', 'name': 'Хуулийн этгээд'}
]
miis_customer_types = [
    {'state': '1', 'name': 'Хувь хүн'},
    {'state': '2', 'name': 'Хуулийн этгээд'},
    {'state': '3', 'name': 'Мэргэшсэн С,Д'},
    {'state': '4', 'name': 'Дамжин өнгөрөх'}
]


class ContractReport(models.Model):
    _name = 'contract.reports'
    
    start_date = fields.Date('Эхлэх огноо', tracking=True, default=lambda self: date.today())
    end_date = fields.Date('Дуусах огноо', tracking=True, default=lambda self: date.today() + relativedelta(months=1))
    user_id = fields.Many2one('res.users', string='Ажилтан', default=lambda self: self.env.user, tracking=True)


    def general_report_excel(self):
        data = {
            'model_id': self.id,
            'start_date': self.start_date,
            'end_date': self.end_date
        }

        return {
            'type': 'ir.actions.report',
            'data': {
                'model': 'contract.reports',
                'options': json.dumps(data, default=date_utils.json_default),
                'output_format': 'xlsx',
                'callback_method': 'get_general_xlsx_report',
                'report_name': 'Дэлгэрэнгүй тайлан',
            },
            'report_type': 'xlsx',
        }
    
    
    def get_general_xlsx_report(self, data, response):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        cell_format = workbook.add_format( {'font_size': '12px', 'align': 'center'})
        head = workbook.add_format({'align': 'center', 'bold': True, 'font_size': '20px'})
        txt = workbook.add_format({'font_size': '10px', 'align': 'center'})

        worksheet.write('E1', 'Дэлгэрэнгүй тайлан', head)
        worksheet.merge_range('I3:J3', f"{data['start_date']} -c {data['end_date']} хүртэл", cell_format)

        
        worksheet.write('A5', 'Д/д', cell_format)
        worksheet.write('B5', 'Харилцагчийн төрөл', cell_format)
        worksheet.write('C5', 'Харилцагчийн овог', cell_format)
        worksheet.write('D5', 'Харилцагчийн нэр', cell_format)
        worksheet.write('E5', 'Регистрийн дугаар', cell_format)
        worksheet.write('F5', 'Утасны дугаар', cell_format)
        worksheet.write('G5', 'Улсын дугаар', cell_format)
        worksheet.write('H5', 'Даатгагч', cell_format)
        worksheet.write('I5', 'Гэрээний дугаар', cell_format)
        worksheet.write('J5', 'Хэрэглэгч', cell_format)
        worksheet.write('K5', 'Даатгалын бүтээгдэхүүн', cell_format)
        worksheet.write('L5', 'Үнэлгээ', cell_format)
        worksheet.write('M5', 'Төлбөр', cell_format)
        worksheet.write('N5', 'Хувь', cell_format)
        worksheet.write('O5', 'Эхлэх огноо', cell_format)
        worksheet.write('P5', 'Дуусах огноо', cell_format)
        worksheet.write('Q5', 'Бүртгэсэн огноо', cell_format)
        worksheet.write('R5', 'Статус', cell_format)
        worksheet.write('S5', 'Даатгалын төрөл', cell_format)

        contracts = self.env['contracts'].search_read(domain=[('create_date', '>=', data['start_date']), ('create_date', '<=', data['end_date']), ('state', '!=', 'draft')],
            fields=field_names, order='start_date asc')
        miis = self.env['miis'].search_read(domain=[('create_date', '>=', data['start_date']), ('create_date', '<=', data['end_date']), ('state', '=', 'done')], 
                fields=field_names, order='start_date asc')
        
        row = 6
        for i, contract in enumerate(contracts):
            worksheet.write(row, 0, i + 1, cell_format)
            for field_index in range(0, 17):
                col = field_index + 1
                if field_names[field_index] == 'state':
                    for state in state_names:
                        if(contract[field_names[field_index]] == state['state']):
                            worksheet.write(row, col, state['name'], cell_format)
                elif field_names[field_index] == 'customer_type':
                    for type in customer_types:
                        if(contract[field_names[field_index]] == type['state']):
                            worksheet.write(row, col, type['name'], cell_format)
                elif field_names[field_index] in ['start_date', 'end_date', 'create_date']:
                    worksheet.write(row, col, contract[field_names[field_index]].strftime('%Y-%m-%d'), txt)
                else:
                    worksheet.write(row, col, contract[field_names[field_index]], cell_format)
            row += 1
        
        row += 2
        for i, contract in enumerate(miis):
            worksheet.write(row, 0, i + 1, cell_format)
            for field_index in range(0, 17):
                col = field_index + 1
                if field_names[field_index] == 'state':
                    worksheet.write(row, col, 'Төлөгдсөн', cell_format)
                elif field_names[field_index] == 'customer_type':
                    for type in miis_customer_types:
                        if(contract[field_names[field_index]] == type['state']):
                            worksheet.write(row, col, type['name'], cell_format)
                elif field_names[field_index] in ['start_date', 'end_date', 'create_date']:
                    worksheet.write(row, col, contract[field_names[field_index]].strftime('%Y-%m-%d'), txt)
                else:
                    worksheet.write(row, col, contract[field_names[field_index]], cell_format)
            row += 1

        
        workbook.close()
        output.seek(0)
        response.stream.write(output.read())
        output.close()


    def reward_report_excel(self):
        data = {
            'model_id': self.id,
            'user_id': self.user_id.id,
            'start_date': self.start_date,
            'end_date': self.end_date
        }

        return {
            'type': 'ir.actions.report',
            'data': {
                'model': 'contract.reports',
                'options': json.dumps(data, default=date_utils.json_default),
                'output_format': 'xlsx',
                'callback_method': 'get_reward_xlsx_report',
                'report_name': 'Урамшууллын тайлан',
            },
            'report_type': 'xlsx',
        }
    

    def get_reward_xlsx_report(self, data, response):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        cell_format = workbook.add_format( {'font_size': '12px', 'align': 'center'})
        head = workbook.add_format({'align': 'center', 'bold': True, 'font_size': '20px'})
        txt = workbook.add_format({'font_size': '10px', 'align': 'center'})

        worksheet.write('E1', 'УРАМШУУЛАЛЫН ТООЦОО', head)
        worksheet.merge_range('I3:J3', f"{data['start_date']} -c {data['end_date']} хүртэл", cell_format)

        worksheet.write('A5', 'Д/д', cell_format)
        worksheet.write('B5', 'Бүртгэсэн огноо', cell_format)
        worksheet.write('C5', 'Ажилтаны нэр', cell_format)
        worksheet.write('D5', 'Гэрээний дугаар', cell_format)
        worksheet.write('E5', 'Хураамж', cell_format)
        worksheet.write('F5', 'Хязгаарлаагүй', cell_format)

        row = 6
        worksheet.write(row, 0, 'Албан журмын даатгал', cell_format)
        row += 1
        miises = self.env['miis'].search_read(domain=[('create_date', '>=', data['start_date']), ('create_date', '<=', data['end_date']), 
                                                    ('state', '=', 'done'), ('user_id', '=', data['user_id'])], 
                fields=field_names_1, order='start_date asc')
        
        for i, contract in enumerate(miises):
            worksheet.write(row, 0, i + 1, cell_format)
            for field_index in range(0, 5):
                col = field_index + 1
                if field_names_1[field_index] == 'create_date':
                    worksheet.write(row, col, contract[field_names_1[field_index]].strftime('%Y-%m-%d'), txt)
                elif(field_names_1[field_index] == 'is_limit'):
                    worksheet.write(row, col, 'Хязгаарласан' if contract[field_names_1[field_index]] else 'Хязгаарлаагүй', cell_format)
                else:
                    worksheet.write(row, col, contract[field_names_1[field_index]], cell_format)
            row += 1


        insurance_types = self.env['insurance.types'].search([])
        for num, insurance_type in enumerate(insurance_types):
            worksheet.write(row, 0, insurance_type.name, cell_format)
            row += 1
            contracts = self.env['contracts'].search_read(domain=[('create_date', '>=', data['start_date']), ('create_date', '<=', data['end_date']), 
                                    ('state', '!=', 'draft'), ('user_id', '=', data['user_id']), ('insurance_type_id', '=', insurance_type.id)],
                                fields=field_names_1, order='start_date asc')
            for i, contract in enumerate(contracts):
                    worksheet.write(row, 0, i + 1, cell_format)
                    for field_index in range(0, 5):
                        col = field_index + 1
                        if field_names_1[field_index] == 'create_date':
                            worksheet.write(row, col, contract[field_names_1[field_index]].strftime('%Y-%m-%d'), txt)
                        elif(field_names_1[field_index] == 'is_limit'):
                            worksheet.write(row, col, 'Хязгаарласан' if contract[field_names_1[field_index]] else 'Хязгаарлаагүй', cell_format)
                        else:
                            worksheet.write(row, col, contract[field_names_1[field_index]], cell_format)
                    row += 1
        
        workbook.close()
        output.seek(0)
        response.stream.write(output.read())
        output.close()


    def financial_commision_report_excel(self):
        data = {
            'model_id': self.id,
            'start_date': self.start_date,
            'end_date': self.end_date
        }

        return {
            'type': 'ir.actions.report',
            'data': {
                'model': 'contract.reports',
                'options': json.dumps(data, default=date_utils.json_default),
                'output_format': 'xlsx',
                'callback_method': 'get_financial_xlsx_report',
                'report_name': 'Санхүүгийн зохицуулах хорооны тайлан',
            },
            'report_type': 'xlsx',
        }
    

    def get_financial_xlsx_report(self, data, response):
        start_date = data['start_date']
        end_date = data['end_date']
        # workbook = load_workbook('broker/broker_contract/static/src/xls/financialReport.xlsx')
        workbook = load_workbook('/home/odoo//www/odoo-digital-zuuch/broker/broker_contract/static/src/xls/financialReport.xlsx')
        contracts = self.env['contracts'].search([('create_date', '>=', data['start_date']), ('create_date', '<=', data['end_date']), ('state', '!=', 'draft')],
                order='insurance_id asc')
        miises = self.env['miis'].search([('create_date', '>=', data['start_date']), ('create_date', '<=', data['end_date']), ('state', '=', 'done')], 
                order='insurance_id asc')
        
        # ------------------------------------ sheet 35 ------------------------------------
        sheet_2 = workbook['35']
        row = 12
        brokers = self.env['res.partner'].search([('broker_type', '=', 'insurance')])
        for i, broker in enumerate(brokers):
            total_valuation_amount = 0
            total_paid_amount = 0
            total_person_paid_amount = 0
            total_fee_amount = 0

            for j, contract in enumerate(contracts):
                if(contract.insurance_id.id == broker.id):
                    total_valuation_amount += contract.valuation
                    total_paid_amount += contract.total_payment
                    if(contract.customer_type == 'person'):
                        total_person_paid_amount += contract.total_payment
                    total_fee_amount += contract.payment_fee_percent

            for j, miis in enumerate(miises):
                if(miis.insurance_id.id == broker.id):
                    total_valuation_amount += miis.valuation
                    total_paid_amount += miis.paid
                    if(miis.customer_type in ['1','4']):
                        total_person_paid_amount += miis.paid
                    total_fee_amount += miis.payment_fee_percent


            sheet_2['B' + str(row)] = broker.name
            sheet_2['D' + str(row)] = total_valuation_amount
            sheet_2['E' + str(row)] = total_person_paid_amount
            sheet_2['J' + str(row)] = (total_paid_amount - total_person_paid_amount)
            sheet_2['K' + str(row)] = total_paid_amount
            sheet_2['L' + str(row)] = total_fee_amount
            row += 1

        # ------------------------------------ sheet 34 ------------------------------------
        sheet = workbook['34']
        data = {
            'car': {
                'valuation': 0, 'paid': 0, 'person_paid': 0, 'count': 0, 'broker_fee': 0, 'row': 14
            },
            'travel': {
                'valuation': 0, 'paid': 0, 'person_paid': 0, 'count': 0, 'broker_fee': 0, 'row': 12
            },
            'miis': {
                'valuation': 0, 'paid': 0, 'person_paid': 0, 'count': 0, 'broker_fee': 0, 'row': 34
            }
        }

        for contract in contracts:
            slug = contract.insurance_type_id.slug
            if slug == 'car':
                key = 'car'
            elif slug in ('travel', 'local_travel'):
                key = 'travel'
            else:
                continue

            data[key]['valuation'] += contract.valuation
            data[key]['paid'] += contract.total_payment
            data[key]['count'] += 1
            data[key]['broker_fee'] += (contract.total_payment * contract.product_id.broker_fee_percent) / 100
            if contract.customer_type == 'person':
                data[key]['person_paid'] += contract.total_payment

        # Process MIIS data
        for miis in miises:
            data['miis']['valuation'] += miis.valuation
            data['miis']['paid'] += miis.paid
            data['miis']['count'] += 1
            data['miis']['broker_fee'] += (miis.paid * miis.product_id.broker_fee_percent) / 100
            if miis.customer_type in ['1', '4']:
                data['miis']['person_paid'] += miis.paid

        # Populate sheet
        for key, values in data.items():
            row = values['row']
            sheet[f'D{row}'] = values['valuation']
            sheet[f'E{row}'] = values['paid']
            sheet[f'F{row}'] = values['person_paid']
            sheet[f'G{row}'] = values['paid'] - values['person_paid']
            sheet[f'H{row}'] = values['count']
            sheet[f'I{row}'] = values['broker_fee']

        # -------------------------------------- sheet 46 ------------------------------------
        sheet_4 = workbook['46']
        data_fields = ['male_18', 'female_18', 'male_19_35', 'female_19_35', 'male_36_55', 'female_36_55', 'male_56', 'female_56']
        counts = {
            'car': {'row': 16, **{field: 0 for field in data_fields}},
            'travel': {'row': 12, **{field: 0 for field in data_fields}},
            # 'local_travel': {'row': 12, **{field: 0 for field in data_fields}},
            'miis': {'row': 31, **{field: 0 for field in data_fields}},
        }

        for j, contract in enumerate(contracts):
            slug = contract.insurance_type_id.slug
            if slug in counts:  # Ensure the slug is valid
                age = int(contract.customer_id.age)
                gender = contract.customer_id.gender
                
                if age <= 18:
                    key = 'male_18' if gender == 'male' else 'female_18'
                elif age <= 35:
                    key = 'male_19_35' if gender == 'male' else 'female_19_35'
                elif age <= 55:
                    key = 'male_36_55' if gender == 'male' else 'female_36_55'
                else:
                    key = 'male_56' if gender == 'male' else 'female_56'
                counts[slug][key] += 1
                
        # for j, miis in enumerate(miises):
        #     age = int(miis.customer_id.age)
        #     gender = miis.customer_id.gender
            
        #     if age <= 18:
        #         key = 'male_18' if gender == 'male' else 'female_18'
        #     elif age <= 35:
        #         key = 'male_19_35' if gender == 'male' else 'female_19_35'
        #     elif age <= 55:
        #         key = 'male_36_55' if gender == 'male' else 'female_36_55'
        #     else:
        #         key = 'male_56' if gender == 'male' else 'female_56'
        #     counts['miis'][key] += 1
        
        columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']  # Corresponding columns
        for counttype in counts:
            nrow = counts[counttype]['row']
            for col, field in zip(columns, data_fields):
                sheet_4[f'{col}{nrow}'] = counts[counttype][field]

        # -------------------------------------- sheet 36 ------------------------------------
        sheet_3 = workbook['36']
        row = 11
        branches = self.env['hr.department'].search([])
        for i, branch in enumerate(branches):
            total_valuation_amount = 0
            total_paid_amount = 0
            total_person_count = 0
            total_count = 0
            total_fee_amount = 0

            for j, contract in enumerate(contracts):
                if(contract.user_id.branch.id == branch.id):
                    total_count += 1
                    total_valuation_amount += contract.valuation
                    total_paid_amount += contract.total_payment
                    if(contract.customer_type == 'person'):
                        total_person_count += 1
                    total_fee_amount += contract.payment_fee_percent

            for j, miis in enumerate(miises):
                if(miis.user_id.branch.id == branch.id):
                    total_count += 1
                    total_valuation_amount += miis.valuation
                    total_paid_amount += miis.paid
                    if(miis.customer_type in ['1','4']):
                        total_person_count += 1
                    total_fee_amount += miis.payment_fee_percent

            sheet_3['A' + str(row)] = i+1
            sheet_3['B' + str(row)] = branch.name
            sheet_3['H' + str(row)] = total_valuation_amount
            sheet_3['I' + str(row)] = total_paid_amount
            sheet_3['J' + str(row)] = total_person_count
            sheet_3['K' + str(row)] = (total_count - total_person_count)
            sheet_3['M' + str(row)] = total_fee_amount

            row += 1


        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        response.stream.write(output.read())
        output.close()
